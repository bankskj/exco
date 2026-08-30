import openpyxl, uuid, sys
from datetime import datetime

def q(s):  # sql string literal
    return "'" + str(s).replace("'", "''") + "'"

def per(dt):  # datetime -> 'YYYY-MM'
    return f"{dt.year}-{dt.month:02d}"

out = []
out.append("-- Exco seed data (generated from Payroll-tracker.xlsx + cashflow.xlsx)")
out.append("DELETE FROM payroll_entries; DELETE FROM employees;")
out.append("DELETE FROM cf_entries; DELETE FROM cf_categories;")

# ---------------- PAYROLL ----------------
wb = openpyxl.load_workbook("Payroll-tracker.xlsx", data_only=True)
ws = wb["HR26"]
# month columns: col5..col18 (headers are dates)
month_cols = []
for c in range(5, ws.max_column+1):
    h = ws.cell(1,c).value
    if isinstance(h, datetime):
        month_cols.append((c, per(h)))
emp_count = 0; pay_count = 0
for r in range(2, 36):  # rows 2..35 = the 34 real employees
    name = ws.cell(r,2).value
    if not name or not str(name).strip():
        continue
    mentor = ws.cell(r,3).value
    # skip rows where the "mentor" cell is numeric (aggregate lines) — none expected in 2..35
    if isinstance(mentor, (int,float)):
        continue
    ctc = ws.cell(r,4).value
    ctc = float(ctc) if isinstance(ctc,(int,float)) else "NULL"
    eid = str(uuid.uuid4()); emp_count += 1
    mentor_sql = q(str(mentor).strip()) if mentor and str(mentor).strip() else "NULL"
    out.append(f"INSERT INTO employees (id,name,mentor,ctc,status,sort_order) VALUES ({q(eid)},{q(str(name).strip())},{mentor_sql},{ctc if ctc!='NULL' else 'NULL'},'active',{emp_count});")
    for (c, pp) in month_cols:
        v = ws.cell(r,c).value
        if isinstance(v,(int,float)):
            out.append(f"INSERT INTO payroll_entries (id,employee_id,period,amount) VALUES ({q(str(uuid.uuid4()))},{q(eid)},{q(pp)},{float(v)});")
            pay_count += 1

# ---------------- CASHFLOW ----------------
wb2 = openpyxl.load_workbook("cashflow.xlsx", data_only=True)
ws2 = wb2["CF26"]
# map 12 columns col2..col13 -> 2025-03 .. 2026-02 (Mar–Feb fiscal year)
cf_periods = []
y, m = 2025, 3
for c in range(2, 14):
    cf_periods.append((c, f"{y}-{m:02d}"))
    m += 1
    if m > 12: m = 1; y += 1

# (sheet_row, kind, group, recurring)
CAT_MAP = [
    (2,  "cost",   "Operating costs",     1),
    (3,  "cost",   "Operating costs",     1),
    (4,  "cost",   "Operating costs",     1),
    (8,  "income", "Recurring income",    1),
    (9,  "income", "Recurring income",    1),
    (10, "income", "Project income",      0),
    (11, "income", "Project income",      0),
    (12, "income", "Recurring income",    1),
    (13, "income", "Project income",      0),
    (14, "income", "Other income",        0),
    (23, "income", "Pipeline (potential)",0),
    (24, "income", "Pipeline (potential)",0),
]
cat_count = 0; cf_entry_count = 0
for order,(row,kind,grp,rec) in enumerate(CAT_MAP, start=1):
    label = ws2.cell(row,1).value
    if not label or not str(label).strip():
        continue
    cid = str(uuid.uuid4()); cat_count += 1
    out.append(f"INSERT INTO cf_categories (id,name,kind,grp,is_recurring,sort_order) VALUES ({q(cid)},{q(str(label).strip())},{q(kind)},{q(grp)},{rec},{order});")
    for (c, pp) in cf_periods:
        v = ws2.cell(row,c).value
        if isinstance(v,(int,float)):
            out.append(f"INSERT INTO cf_entries (id,category_id,period,amount,status) VALUES ({q(str(uuid.uuid4()))},{q(cid)},{q(pp)},{float(v)},'actual');")
            cf_entry_count += 1

# opening balance from r19 col2 (Cash balance)
ob = ws2.cell(19,2).value
ob = float(ob) if isinstance(ob,(int,float)) else 1000000.0
out.append(f"INSERT INTO app_meta (key,value,updated_at) VALUES ('cf_opening_balance',{q(ob)},datetime('now')) ON CONFLICT(key) DO UPDATE SET value=excluded.value;")
out.append("INSERT INTO app_meta (key,value,updated_at) VALUES ('cf_opening_period','2025-03',datetime('now')) ON CONFLICT(key) DO UPDATE SET value=excluded.value;")
out.append("INSERT INTO app_meta (key,value,updated_at) VALUES ('cf_actuals_through','2026-02',datetime('now')) ON CONFLICT(key) DO UPDATE SET value=excluded.value;")

with open(f"{sys.argv[1]}/seed.sql","w") as f:
    f.write("\n".join(out) + "\n")

print(f"employees={emp_count} payroll_entries={pay_count}")
print(f"cf_categories={cat_count} cf_entries={cf_entry_count}")
print(f"opening_balance={ob}")
print("payroll months:", [p for _,p in month_cols])
print("cashflow months:", [p for _,p in cf_periods])
