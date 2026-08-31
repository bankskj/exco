#!/usr/bin/env python3
"""Generate hr_seed.sql from an HR export (company-users xlsx, 'report' sheet).
Usage: python3 scripts/import_hr.py <xlsx> <outdir>
Columns: Employee | Email | Position | Team | Direct manager | Employee # | Start date | Last working day | ..."""
import openpyxl, uuid, sys
from datetime import datetime

def q(s): return "'" + str(s).replace("'", "''") + "'"
def d(v):
    if isinstance(v, datetime): return v.strftime("%Y-%m-%d")
    s = str(v).strip() if v else ""
    return s if len(s) == 10 and s[4] == "-" else None

wb = openpyxl.load_workbook(sys.argv[1], data_only=True)
ws = wb["report"]
out = ["DELETE FROM hr_notes; DELETE FROM hr_employees;"]
n = 0
for r in range(2, ws.max_row + 1):
    name = ws.cell(r, 1).value
    if not name or not str(name).strip(): continue
    vals = {
        "name": " ".join(str(name).split()),
        "email": ws.cell(r, 2).value, "position": ws.cell(r, 3).value,
        "team": ws.cell(r, 4).value, "manager": ws.cell(r, 5).value,
        "employee_no": ws.cell(r, 6).value,
    }
    start, end = d(ws.cell(r, 7).value), d(ws.cell(r, 8).value)
    cols = ", ".join(q(str(v).strip()) if v and str(v).strip() else "NULL"
                     for v in [vals["email"], vals["position"], vals["team"], vals["manager"], vals["employee_no"]])
    out.append(
        f"INSERT INTO hr_employees (id,name,email,position,team,manager,employee_no,start_date,end_date) "
        f"VALUES ({q(str(uuid.uuid4()))},{q(vals['name'])},{cols},"
        f"{q(start) if start else 'NULL'},{q(end) if end else 'NULL'});")
    n += 1
open(f"{sys.argv[2]}/hr_seed.sql", "w").write("\n".join(out) + "\n")
print(f"hr_seed.sql written: {n} employees")
